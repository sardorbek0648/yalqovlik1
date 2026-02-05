# -*- coding: utf-8 -*-
"""
SHOP BOT — FULL SINGLE-FILE VERSION
Author: ChatGPT
Notes:
- Requires: python-telegram-bot v20+, pymysql, openpyxl
- Configure TOKEN, DB creds below.
"""

import os
import logging
from datetime import datetime
import pymysql
from openpyxl import Workbook, load_workbook

from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ContextTypes, ConversationHandler, filters
)

# ===================== LOGGING =====================
logging.basicConfig(
    filename="error.log",
    level=logging.ERROR,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

def log_error(e: Exception):
    logging.exception(e)

# ===================== CONFIG =====================
TOKEN = "8539404211:AAHDyn8jngmul6gO9pqTJP9Vypx9ZFVNtAg"
PRIMARY_ADMIN_ID = 5788278697

DB_HOST = "localhost"
DB_USER = "root"          # consider dedicated user
DB_PASS = "sardorbek06."
DB_NAME = "shop_bot"

# ===================== STATES =====================
(
    # Phone add
    ADD_PHONE_CHOICE,
    ADD_PHONE_MANUAL,
    ADD_PHONE_EXCEL,

    # Sell / Return
    SELL_INPUT_ID,
    SELL_CONFIRM,
    RETURN_INPUT_ID,

    # Admin add
    ADD_ADMIN_ID,
    ADD_ADMIN_USERNAME,

    # Seller add
    ADD_SELLER_ID,
    ADD_SELLER_USERNAME,

    # Remove admin/seller
    REMOVE_PICK
) = range(11)

# ===================== DATABASE =====================
def get_db(no_db: bool = False):
    return pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASS,
        database=None if no_db else DB_NAME,
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True
    )

def init_db():
    try:
        db = get_db(no_db=True)
        with db.cursor() as c:
            c.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME}")
        db.close()

        db = get_db()
        with db.cursor() as c:
            c.execute("""
            CREATE TABLE IF NOT EXISTS users (
                telegram_id BIGINT PRIMARY KEY,
                username VARCHAR(100),
                role ENUM('admin','seller') NOT NULL
            )""")
            c.execute("""
            CREATE TABLE IF NOT EXISTS phones (
                phone_id VARCHAR(50) PRIMARY KEY,
                phone_name VARCHAR(255),
                quantity INT DEFAULT 0
            )""")
            c.execute("""
            CREATE TABLE IF NOT EXISTS sales (
                sale_no INT AUTO_INCREMENT PRIMARY KEY,
                phone_id VARCHAR(50),
                phone_name VARCHAR(255),
                remaining_qty INT,
                sale_type ENUM('SOTILDI','DOKONGA_BERILDI'),
                sale_date DATE,
                sale_time TIME
            )""")
        db.close()
    except Exception as e:
        log_error(e)

def get_role(uid: int):
    if uid == PRIMARY_ADMIN_ID:
        return "primary"
    try:
        db = get_db()
        with db.cursor() as c:
            c.execute("SELECT role FROM users WHERE telegram_id=%s", (uid,))
            r = c.fetchone()
        return r["role"] if r else None
    except Exception:
        return None

def is_int(s: str) -> bool:
    return s.isdigit()

# ===================== MENUS =====================
def menu_primary():
    return ReplyKeyboardMarkup([
        ["📦 Telefon qo‘shish", "📊 Qoldiqni ko‘rish"],
        ["📥 Excel hisobot"],
        ["👤 Admin qo‘shish", "🗑 Adminni olib tashlash"],
        ["👷 Sotuvchi qo‘shish", "🗑 Sotuvchini olib tashlash"],
        ["🧹 Do‘kondagi telefonlarni tozalash"],
        ["📄 Xatoliklarni ko‘rish"]
    ], resize_keyboard=True)

def menu_admin():
    return ReplyKeyboardMarkup([
        ["📦 Telefon qo‘shish", "📊 Qoldiqni ko‘rish"],
        ["📥 Excel hisobot"]
    ], resize_keyboard=True)

def menu_seller():
    return ReplyKeyboardMarkup([
        ["📊 Telefon qoldig'ini tekshirish"],
        ["✅ Telefon sotish"],
        ["🔄 Do‘konga berish"]
    ], resize_keyboard=True)

async def go_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    role = get_role(update.effective_user.id)
    if role == "primary":
        await update.message.reply_text("📌 Panel:", reply_markup=menu_primary())
    elif role == "admin":
        await update.message.reply_text("📌 Panel:", reply_markup=menu_admin())
    elif role == "seller":
        await update.message.reply_text("📌 Panel:", reply_markup=menu_seller())
    else:
        await update.message.reply_text(
            f"❌ Siz botdan foydalana olmaysiz\nTelegram ID: {update.effective_user.id}"
        )

# ===================== START =====================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await go_menu(update, context)

# ===================== STOCK =====================
async def show_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        db = get_db()
        with db.cursor() as c:
            c.execute("SELECT * FROM phones WHERE quantity>0")
            rows = c.fetchall()
        if not rows:
            await update.message.reply_text("📭 Telefon yo‘q")
            await update.message.reply_text("📊 Umumiy: 0")
            return
        total = 0
        msg = "📦 Mavjud telefonlar:\n\n"
        for r in rows:
            msg += f"{r['phone_id']} | {r['phone_name']} | {r['quantity']}\n"
            total += r["quantity"]
        await update.message.reply_text(msg)
        await update.message.reply_text(f"📊 Umumiy: {total}")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")

# ===================== PHONE ADD =====================
async def add_phone_start(update, context):
    await update.message.reply_text(
        "📦 Telefon qo‘shish usulini tanlang:",
        reply_markup=ReplyKeyboardMarkup([
            ["➕ Qo‘lda qo‘shish", "📂 Excel orqali qo‘shish"],
            ["⬅️ Orqaga"]
        ], resize_keyboard=True)
    )
    return ADD_PHONE_CHOICE

async def add_phone_choice(update, context):
    t = update.message.text
    if t == "➕ Qo‘lda qo‘shish":
        await update.message.reply_text("ID,Nomi,Miqdor\nMasalan: 501,Oppo A3,5", reply_markup=ReplyKeyboardRemove())
        return ADD_PHONE_MANUAL
    if t == "📂 Excel orqali qo‘shish":
        await update.message.reply_text("📂 Excel (.xlsx) yuboring", reply_markup=ReplyKeyboardRemove())
        return ADD_PHONE_EXCEL
    await go_menu(update, context)
    return ConversationHandler.END

async def add_phone_manual(update, context):
    try:
        parts = [p.strip() for p in update.message.text.split(",")]
        if len(parts) != 3 or not parts[2].isdigit():
            await update.message.reply_text("❌ Format xato")
            return ConversationHandler.END
        pid, name, qty = parts
        qty = int(qty)
        db = get_db()
        with db.cursor() as c:
            c.execute("SELECT quantity FROM phones WHERE phone_id=%s", (pid,))
            r = c.fetchone()
            if r:
                c.execute("UPDATE phones SET quantity=quantity+%s WHERE phone_id=%s", (qty, pid))
            else:
                c.execute("INSERT INTO phones VALUES (%s,%s,%s)", (pid, name, qty))
        await update.message.reply_text("✅ Telefon qo‘shildi")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")
    await go_menu(update, context)
    return ConversationHandler.END

async def add_phone_excel(update, context):
    try:
        doc = update.message.document
        if not doc or not doc.file_name.endswith(".xlsx"):
            await update.message.reply_text("❌ Faqat .xlsx")
            return ADD_PHONE_EXCEL
        file = await doc.get_file()
        path = f"temp_{doc.file_name}"
        await file.download_to_drive(path)

        wb = load_workbook(path)
        ws = wb.active
        first = ws.cell(1, 1).value
        start_row = 1 if str(first).isdigit() else 2

        added = 0
        db = get_db()
        with db.cursor() as c:
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                try:
                    pid, name, qty = row[:3]
                    if not pid or not name or qty is None:
                        continue
                    qty = int(qty)
                    c.execute("SELECT quantity FROM phones WHERE phone_id=%s", (pid,))
                    r = c.fetchone()
                    if r:
                        c.execute("UPDATE phones SET quantity=quantity+%s WHERE phone_id=%s", (qty, pid))
                    else:
                        c.execute("INSERT INTO phones VALUES (%s,%s,%s)", (pid, name, qty))
                    added += 1
                except Exception as e:
                    log_error(e)
        os.remove(path)
        await update.message.reply_text(f"✅ Excel orqali {added} ta telefon qo‘shildi")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Excel xatosi")
    await go_menu(update, context)
    return ConversationHandler.END

# ===================== SELL / RETURN =====================
async def sell_start(update, context):
    context.user_data.clear()
    context.user_data["mode"] = "SOTILDI"
    await update.message.reply_text("⚠️ MIJOZGA SOTILADI\n📱 Telefon ID ni kiriting:", reply_markup=ReplyKeyboardRemove())
    return SELL_INPUT_ID

async def return_start(update, context):
    context.user_data.clear()
    context.user_data["mode"] = "DOKONGA_BERILDI"
    await update.message.reply_text("⚠️ DO‘KONGA BERILADI\n📱 Telefon ID ni kiriting:", reply_markup=ReplyKeyboardRemove())
    return RETURN_INPUT_ID

async def process_id(update, context):
    pid = update.message.text.strip()
    if not pid.isdigit():
        await update.message.reply_text("❌ ID raqam bo‘lishi kerak")
        return ConversationHandler.END
    try:
        db = get_db()
        with db.cursor() as c:
            c.execute("SELECT * FROM phones WHERE phone_id=%s", (pid,))
            phone = c.fetchone()
        if not phone or phone["quantity"] <= 0:
            await update.message.reply_text("❌ Telefon topilmadi yoki tugagan")
            return ConversationHandler.END
        context.user_data["phone"] = phone
        kb = ReplyKeyboardMarkup([["✅ Tasdiqlash", "❌ Bekor qilish"]], resize_keyboard=True)
        await update.message.reply_text(
            f"📱 {phone['phone_name']}\n🆔 {phone['phone_id']}\n📦 Qoldiq: {phone['quantity']}\nHolat: {context.user_data['mode']}",
            reply_markup=kb
        )
        return SELL_CONFIRM
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")
        return ConversationHandler.END

async def confirm_action(update, context):
    if update.message.text != "✅ Tasdiqlash":
        await update.message.reply_text("❌ Bekor qilindi")
        await go_menu(update, context)
        return ConversationHandler.END
    try:
        phone = context.user_data["phone"]
        mode = context.user_data["mode"]
        db = get_db()
        with db.cursor() as c:
            new_q = phone["quantity"] - 1
            c.execute("UPDATE phones SET quantity=%s WHERE phone_id=%s", (new_q, phone["phone_id"]))
            now = datetime.now()
            c.execute("""
                INSERT INTO sales
                (phone_id, phone_name, remaining_qty, sale_type, sale_date, sale_time)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (phone["phone_id"], phone["phone_name"], new_q, mode, now.date(), now.strftime("%H:%M")))
        await update.message.reply_text("✅ Amal bajarildi")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")
    await go_menu(update, context)
    return ConversationHandler.END

# ===================== ADMIN CRUD =====================
async def add_admin_start(update, context):
    if update.effective_user.id != PRIMARY_ADMIN_ID:
        await update.message.reply_text("❌ Ruxsat yo‘q")
        return ConversationHandler.END
    await update.message.reply_text("👤 Admin Telegram ID:")
    return ADD_ADMIN_ID

async def add_admin_id(update, context):
    if not update.message.text.isdigit():
        await update.message.reply_text("❌ Faqat raqam")
        return ADD_ADMIN_ID
    context.user_data["uid"] = update.message.text
    await update.message.reply_text("👤 Admin username:")
    return ADD_ADMIN_USERNAME

async def add_admin_username(update, context):
    try:
        db = get_db()
        with db.cursor() as c:
            c.execute("REPLACE INTO users VALUES (%s,%s,'admin')", (context.user_data["uid"], update.message.text))
        await update.message.reply_text("✅ Admin qo‘shildi")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")
    await go_menu(update, context)
    return ConversationHandler.END

async def add_seller_start(update, context):
    if update.effective_user.id != PRIMARY_ADMIN_ID:
        await update.message.reply_text("❌ Ruxsat yo‘q")
        return ConversationHandler.END
    await update.message.reply_text("👷 Sotuvchi Telegram ID:")
    return ADD_SELLER_ID

async def add_seller_id(update, context):
    if not update.message.text.isdigit():
        await update.message.reply_text("❌ Faqat raqam")
        return ADD_SELLER_ID
    context.user_data["uid"] = update.message.text
    await update.message.reply_text("👷 Sotuvchi username:")
    return ADD_SELLER_USERNAME

async def add_seller_username(update, context):
    try:
        db = get_db()
        with db.cursor() as c:
            c.execute("REPLACE INTO users VALUES (%s,%s,'seller')", (context.user_data["uid"], update.message.text))
        await update.message.reply_text("✅ Sotuvchi qo‘shildi")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")
    await go_menu(update, context)
    return ConversationHandler.END

async def remove_pick(update, context):
    try:
        role = update.message.text
        db = get_db()
        with db.cursor() as c:
            c.execute("SELECT telegram_id, username FROM users WHERE role=%s", (role,))
            rows = c.fetchall()
        if not rows:
            await update.message.reply_text("📭 Ro‘yxat bo‘sh")
            await go_menu(update, context)
            return ConversationHandler.END
        kb = [[f"❌ @{r['username']} | {r['telegram_id']}"] for r in rows]
        await update.message.reply_text("Tanlang:", reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True))
        context.user_data["remove_role"] = role
        return REMOVE_PICK
    except Exception as e:
        log_error(e)
        await go_menu(update, context)
        return ConversationHandler.END

async def remove_confirm(update, context):
    try:
        uid = update.message.text.split("|")[1].strip()
        role = context.user_data.get("remove_role")
        db = get_db()
        with db.cursor() as c:
            c.execute("DELETE FROM users WHERE telegram_id=%s AND role=%s", (uid, role))
        await update.message.reply_text("✅ O‘chirildi")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")
    await go_menu(update, context)
    return ConversationHandler.END

# ===================== UTIL =====================
async def clear_store(update, context):
    if update.effective_user.id != PRIMARY_ADMIN_ID:
        await update.message.reply_text("❌ Ruxsat yo‘q")
        return
    try:
        db = get_db()
        with db.cursor() as c:
            c.execute("DELETE FROM phones")
        await update.message.reply_text("🧹 Barcha sotilmagan telefonlar o‘chirildi")
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Xato")
    await go_menu(update, context)

async def export_excel(update, context):
    try:
        db = get_db()
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.append(["ID","Nomi","Qoldiq"])
        with db.cursor() as c:
            c.execute("SELECT * FROM phones")
            for r in c.fetchall():
                ws1.append([r["phone_id"], r["phone_name"], r["quantity"]])
        wb1.save("sotilmagan.xlsx")
        await update.message.reply_document(open("sotilmagan.xlsx","rb"))

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["№","ID","Nomi","Qoldiq","Sana","Vaqt","Holat"])
        with db.cursor() as c:
            c.execute("SELECT * FROM sales ORDER BY sale_no")
            for r in c.fetchall():
                ws2.append([r["sale_no"], r["phone_id"], r["phone_name"], r["remaining_qty"], r["sale_date"], r["sale_time"], r["sale_type"]])
        wb2.save("sotilgan.xlsx")
        await update.message.reply_document(open("sotilgan.xlsx","rb"))
    except Exception as e:
        log_error(e)
        await update.message.reply_text("❌ Excel xatosi")

async def send_log(update, context):
    try:
        await update.message.reply_document(open("error.log","rb"))
    except Exception:
        await update.message.reply_text("📄 Log yo‘q")

# ===================== MAIN =====================
def main():
    init_db()
    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))

    app.add_handler(ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📦 Telefon qo‘shish$"), add_phone_start)],
        states={
            ADD_PHONE_CHOICE: [MessageHandler(filters.TEXT, add_phone_choice)],
            ADD_PHONE_MANUAL: [MessageHandler(filters.TEXT, add_phone_manual)],
            ADD_PHONE_EXCEL: [MessageHandler(filters.Document.ALL, add_phone_excel)],
        },
        fallbacks=[]
    ))

    app.add_handler(MessageHandler(filters.Regex("Qoldiq"), show_stock))
    app.add_handler(MessageHandler(filters.Regex("^📊 Telefon qoldig'ini tekshirish$"), show_stock))

    app.add_handler(ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^✅ Telefon sotish$"), sell_start)],
        states={
            SELL_INPUT_ID: [MessageHandler(filters.TEXT, process_id)],
            SELL_CONFIRM: [MessageHandler(filters.TEXT, confirm_action)],
        },
        fallbacks=[]
    ))

    app.add_handler(ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^🔄 Do‘konga berish$"), return_start)],
        states={
            RETURN_INPUT_ID: [MessageHandler(filters.TEXT, process_id)],
            SELL_CONFIRM: [MessageHandler(filters.TEXT, confirm_action)],
        },
        fallbacks=[]
    ))

    app.add_handler(ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^👤 Admin qo‘shish$"), add_admin_start)],
        states={
            ADD_ADMIN_ID: [MessageHandler(filters.TEXT, add_admin_id)],
            ADD_ADMIN_USERNAME: [MessageHandler(filters.TEXT, add_admin_username)],
        },
        fallbacks=[]
    ))

    app.add_handler(ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^👷 Sotuvchi qo‘shish$"), add_seller_start)],
        states={
            ADD_SELLER_ID: [MessageHandler(filters.TEXT, add_seller_id)],
            ADD_SELLER_USERNAME: [MessageHandler(filters.TEXT, add_seller_username)],
        },
        fallbacks=[]
    ))

    app.add_handler(MessageHandler(filters.Regex("^🗑 Adminni olib tashlash$"), lambda u,c: remove_pick(u,c)))
    app.add_handler(MessageHandler(filters.Regex("^🗑 Sotuvchini olib tashlash$"), lambda u,c: remove_pick(u,c)))
    app.add_handler(MessageHandler(filters.Regex("^❌ @"), remove_confirm))

    app.add_handler(MessageHandler(filters.Regex("^🧹 Do‘kondagi telefonlarni tozalash$"), clear_store))
    app.add_handler(MessageHandler(filters.Regex("^📥 Excel hisobot$"), export_excel))
    app.add_handler(MessageHandler(filters.Regex("^📄 Xatoliklarni ko‘rish$"), send_log))

    app.run_polling()

if __name__ == "__main__":
    main()
